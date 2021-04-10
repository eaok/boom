#!/usr/bin/env python3
#-*- coding: utf-8 -*-
'Convert excel files to database files.'
__author__ = "wangpan"

import os
import sqlite3
import openpyxl

class ExcelToDatabase(object):
    def __init__(self, excel_path, sheet_name, data_file):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.data_file = data_file

        if os.path.exists(self.data_file):
            os.remove(self.data_file)

    def create_db(self):
        conn = sqlite3.connect(self.data_file)
        print("Opened database successfully!")

        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS beatcolor (
            ChineseID         INT,
            EnglishID         INT,
            ChineseRules      STRING,
            EnglishRules      STRING,
            VietnameseRules   STRING,
            RussianRules      STRING,
            RemoteControlType INT,
            SettingPoint      INT,
            AppointPosition   INT,
            BeatColorType     INT,
            ApplicableVersion INT,
            XParameters       STRING,
            YParameters       STRING,
            ZParameters       STRING
        );''')
        print("Table created successfully")
        conn.commit()
        conn.close()

    def __get_string_part(self, value, flag_type, flag_first):
        if flag_type:
            if flag_first:
                if not isinstance(value, int):
                    ret = "null"
                else:
                    ret = str(value)
            else:
                if not isinstance(value, int):
                    ret = ", null"
                else:
                    ret = ", " + str(value)

        else:
            if not isinstance(value, str):
                ret = ", null"
            else:
                ret = ", \"" + value + "\""

        return ret

    def read_excel(self):
        book = openpyxl.load_workbook(self.excel_path)
        sheet = book[self.sheet_name]

        conn = sqlite3.connect(self.data_file)
        c = conn.cursor()
        for row in range(2, sheet.max_row + 1, 1):
            ChineseID = self.__get_string_part(sheet.cell(row, column=1).value, True, True)
            EnglishID = self.__get_string_part(sheet.cell(row, column=2).value, True, False)
            ChineseRules = self.__get_string_part(sheet.cell(row, column=3).value, False, False)
            EnglishRules = self.__get_string_part(sheet.cell(row, column=4).value, False, False)
            VietnameseRules = self.__get_string_part(sheet.cell(row, column=5).value, False, False)
            RussianRules = self.__get_string_part(sheet.cell(row, column=6).value, False, False)
            RemoteControlType = self.__get_string_part(sheet.cell(row, column=7).value, True, False)
            SettingPoint = self.__get_string_part(sheet.cell(row, column=8).value, True, False)
            AppointPosition = self.__get_string_part(sheet.cell(row, column=9).value, True, False)
            BeatColorType = self.__get_string_part(sheet.cell(row, column=10).value, True, False)
            ApplicableVersion = self.__get_string_part(sheet.cell(row, column=11).value, True, False)
            XParameters = self.__get_string_part(sheet.cell(row, column=12).value, False, False)
            YParameters = self.__get_string_part(sheet.cell(row, column=13).value, False, False)
            ZParameters = self.__get_string_part(sheet.cell(row, column=14).value, False, False)

            sql = "INSERT INTO beatcolor ( \
                    ChineseID, EnglishID, ChineseRules, EnglishRules, VietnameseRules, \
                    RussianRules, RemoteControlType, SettingPoint, AppointPosition, BeatColorType, \
                    ApplicableVersion, XParameters, YParameters, ZParameters) \
                    VALUES (" + ChineseID + EnglishID + ChineseRules + EnglishRules + \
                    VietnameseRules + RussianRules + RemoteControlType + SettingPoint + \
                    BeatColorType + AppointPosition + ApplicableVersion + XParameters + \
                    YParameters + ZParameters + ");"
            print(sql)
            c.execute(sql)

        conn.commit()
        conn.close()

if __name__ == '__main__':
    excel_path = "./data.xlsx"
    sheet_name = "beatcolor"
    data_file = "./data.db"

    obj = ExcelToDatabase(excel_path, sheet_name, data_file)
    obj.create_db()
    obj.read_excel()
